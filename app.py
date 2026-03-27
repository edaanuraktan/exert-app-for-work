import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tempfile

# ---------------------------------
# Streamlit Ayarları
# ---------------------------------
st.set_page_config(page_title="EXERT", layout="centered")

st.title("📊 EXERT")
st.subheader("Akıllı Excel-Insert Karşılaştırma Aracı")

# ---------------------------------
# Renkler
# ---------------------------------
GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

# ---------------------------------
# Karşılaştırma Motoru
# ---------------------------------
def compare_excels(old_file, new_file):
    df_old = pd.read_excel(old_file, header=None)
    df_new = pd.read_excel(new_file, header=None)

    # İlk hücre = satır anahtarı
    old_map = {
        str(row[0]): row
        for _, row in df_old.iterrows()
        if pd.notna(row[0])
    }

    new_map = {
        str(row[0]): row
        for _, row in df_new.iterrows()
        if pd.notna(row[0])
    }

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        output_path = tmp.name

    df_new.to_excel(output_path, index=False, header=False)

    wb = load_workbook(output_path)
    ws = wb.active

    # Yeni ve güncellenen satırlar
    for row_idx, row in df_new.iterrows():
        key = str(row[0])
        excel_row = row_idx + 1

        if key not in old_map:
            # Eğer satır tamamen boşsa hiçbir şey yapma
            if row.isna().all():
                continue

            # Yeni satır → yeşil
            for col in range(1, ws.max_column + 1):
                ws.cell(row=excel_row, column=col).fill = GREEN
        else:
            # Aynı satır → hücresel karşılaştır
            old_row = old_map[key]

            for col_idx in range(1, len(row)):
                new_val = row[col_idx]
                old_val = old_row[col_idx] if col_idx < len(old_row) else None

                if pd.isna(new_val) and pd.isna(old_val):
                    continue

                if str(new_val) != str(old_val):
                    ws.cell(row=excel_row, column=col_idx + 1).fill = YELLOW

    # Silinen satırlar (alta ekle)
    deleted_keys = set(old_map.keys()) - set(new_map.keys())
    start_row = ws.max_row + 2

    if deleted_keys:
        ws.cell(row=start_row - 1, column=1, value="SİLİNEN SATIRLAR")

        for key in deleted_keys:
            row = old_map[key]
            for col_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=start_row, column=col_idx, value=val)
                cell.fill = RED
            start_row += 1

    wb.save(output_path)
    return output_path

# ---------------------------------
# Arayüz
# ---------------------------------
old_file = st.file_uploader("📁 Eski Excel", type=["xlsx"])
new_file = st.file_uploader("📁 Yeni Excel (esas alınır)", type=["xlsx"])

if old_file and new_file:
    st.success("Dosyalar yüklendi")

    if st.button("🔍 Karşılaştır"):
        with st.spinner("Hücresel analiz yapılıyor..."):
            try:
                output = compare_excels(old_file, new_file)

                with open(output, "rb") as f:
                    st.download_button(
                        label="⬇️ Karşılaştırılmış Excel’i İndir",
                        data=f,
                        file_name="EXERT_Comparison.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                st.success("Karşılaştırma tamamlandı")

            except Exception as e:
                st.error(f"Hata: {e}")
